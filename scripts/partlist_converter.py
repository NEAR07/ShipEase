import os
import re
import openpyxl
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

# Modified process_file function to handle profile type and material
def process_file(file_path):
    profile_types = []
    materials = []
    bar_numbers = []

    with open(file_path, 'r') as file:
        current_profile = None
        current_material = None
        for line in file:
            line = line.replace('|', '').strip()

            # Detect Profile type
            profile_match = re.search(r"Profile type\s*:\s*(\S+)", line)
            if profile_match:
                current_profile = profile_match.group(1)

            # Detect Material
            material_match = re.search(r"Material\s*:\s*(\S+)", line)
            if material_match:
                current_material = material_match.group(1)

            # Detect Bar number and append Profile type, Material, and Bar number
            bar_number_match = re.search(r"Bar number\s*:\s*(\d+)", line)
            if bar_number_match and current_profile and current_material:
                profile_types.append(current_profile)
                materials.append(current_material)
                bar_numbers.append(int(bar_number_match.group(1)))

    # Aggregate by Profile type and Material
    profile_material_summary = {}
    for profile, material, bar in zip(profile_types, materials, bar_numbers):
        key = (profile, material)
        if key not in profile_material_summary:
            profile_material_summary[key] = []
        profile_material_summary[key].append(bar)

    # Create the resulting DataFrame with aggregated counts
    profile_type_resumes = []
    material_resumes = []
    bar_number_resumes = []
    for (profile, material), bars in profile_material_summary.items():
        profile_type_resumes.append(profile)
        material_resumes.append(material)
        bar_number_resumes.append(len(bars))  # Count occurrences for each profile-material combination

    df = pd.DataFrame({
        "Profile type resume": profile_type_resumes,
        "Material resume": material_resumes,
        "Bar number resume": bar_number_resumes
    })
    
    return df

# Function to split the Profile type into separate columns
def split_profile_type(profile_type):
    pattern = r'[A-Za-z]+|\d+|X'
    matches = re.findall(pattern, profile_type)
    
    # Ensure the result has 8 elements with default values
    default_values = ['X', '0', 'X', '0', 'X', '0', 'X', '0']
    matches.extend(default_values[len(matches):])
    
    return matches[:8]

# Function to parse the .list file
def parse_list_file(file_path):
    data = {
        "Profile type": [],
        "Bar-codenr": [],
        "Length bar": [],
        "Material": [],
        "Bar number": [],
        "Total length": [],
        "Scrap-iron": [],
        "Part": [],
        "Cut off Length": []
    }

    header_data = {
        "Object": "",
        "Block": "",
        "Date": ""
    }

    with open(file_path, 'r') as file:
        part_section = False
        common_data = {}
        for line in file:
            line = line.replace('|', '').strip()

            object_match = re.search(r"Object:\s*(\d+)", line)
            block_match = re.search(r"Block:\s*(\d+)", line)
            date_match = re.search(r"Date:\s*(\S+)", line)

            if object_match:
                header_data["Object"] = object_match.group(1)
            if block_match:
                header_data["Block"] = block_match.group(1)
            if date_match:
                header_data["Date"] = date_match.group(1)

            if re.match(r"^Part\s+Cut off Length", line):
                part_section = True
                continue

            if part_section and re.match(r"^\d+\s+\d+", line):
                parts = line.split()
                if len(parts) >= 2:
                    data["Part"].append(parts[0])
                    data["Cut off Length"].append(parts[1])
                    for key in common_data:
                        data[key].append(common_data[key])
                continue

            profile_match = re.search(r"Profile type\s*:\s*(.*)", line)
            barcode_match = re.search(r"Bar-codenr\s*:\s*(.*)", line)
            length_bar_match = re.search(r"Length bar\s*:\s*(.*)", line)
            material_match = re.search(r"Material\s*:\s*(.*)", line)
            bar_number_match = re.search(r"Bar number\s*:\s*(.*)", line)
            total_length_match = re.search(r"Total length\s*:\s*(.*)", line)
            scrap_iron_match = re.search(r"Scrap-iron\s*:\s*(.*)", line)

            if profile_match:
                common_data["Profile type"] = profile_match.group(1).split()[0].strip()
            if barcode_match:
                common_data["Bar-codenr"] = barcode_match.group(1).strip()
            if length_bar_match:
                common_data["Length bar"] = length_bar_match.group(1).split()[0].strip()
            if material_match:
                common_data["Material"] = material_match.group(1).split()[0].strip()
            if bar_number_match:
                common_data["Bar number"] = bar_number_match.group(1).strip()
            if total_length_match:
                common_data["Total length"] = total_length_match.group(1).strip()
            if scrap_iron_match:
                common_data["Scrap-iron"] = scrap_iron_match.group(1).strip()

    return data, header_data

# Function to parse the .lst file
def parse_lst_file(lst_file_path):
    lst_data = {}
    with open(lst_file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith('*'):
                parts = line.split('|')
                if len(parts) >= 6:
                    bcode = parts[0].strip('* ')
                    length = parts[5].strip()
                    if length not in lst_data:
                        lst_data[length] = []
                    lst_data[length].append(bcode)
    return lst_data

def parse_lst_file_with_part(lst_file_path):
    lst_data_with_part = {}
    with open(lst_file_path, 'r') as file:
        for line in file:
            line = line.strip()
            if line.startswith('*'):
                parts = line.split('|')
                if len(parts) >= 6:
                    bcode = parts[0].strip('* ')
                    length = parts[5].strip()
                    bcode_parts = bcode.split('-')
                    if len(bcode_parts) >= 3:  # Changed to support 3 or more segments
                        part = bcode_parts[-1]  # Use the last segment as part
                        key = (length, part)
                        if key not in lst_data_with_part:
                            lst_data_with_part[key] = []
                        lst_data_with_part[key].append(bcode)
    return lst_data_with_part

# Function to load CSV data
def load_csv_data(csv_file_path):
    csv_data = []
    with open(csv_file_path, newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=';', quotechar='"')
        for row in reader:
            csv_data.append(row)
    return csv_data

# Function to convert specific columns to numbers
def convert_to_number(ws, row):
    columns_to_convert = [2, 3, 5, 6, 7, 9, 14, 13, 15, 16, 17, 19]
    for col in columns_to_convert:
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            try:
                cell.value = float(cell.value)
            except ValueError:
                pass

# Function to convert barcode columns
def convert_to_number_barcode(ws, row):
    columns_to_convert = [2]
    for col in columns_to_convert:
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            try:
                value_as_int = int(float(cell.value))
                cell.value = str(value_as_int)
                cell.number_format = '@'
            except ValueError:
                pass


def convert_list_to_xlsx(list_file_path, lst_file_path, resume_data, csv_file_path, output_file_path):
    data, header_data = parse_list_file(list_file_path)
    lst_data = parse_lst_file(lst_file_path)
    lst_data_with_part = parse_lst_file_with_part(lst_file_path)
    csv_data = load_csv_data(csv_file_path)

    wb = Workbook()
    ws = wb.active

    ws.append(["PROJECT =", header_data["Object"], "BLOCK =", header_data["Block"], "DATE =", header_data["Date"], "", "", "", "", "", "RESUME"])
    headers = [
        "PROFILE TYPE", "BAR-CODE", "LENGTH BAR", "MAT", "BAR NUMBER",
        "TOT LENGTH", "SCRAP IRON", "PART NAME", "Cut off Length", "", "",
        "BAR-CODE", "Profile Type", "Height", "", "Width", "", "Thick1", "", "Thick2", "MAT2", "Bar number resume"
    ]
    ws.append(headers)

    bold_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_font
    for cell in ws[2]:
        cell.font = bold_font

    # Set column widths
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 9
    ws.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 21
    ws.column_dimensions[openpyxl.utils.get_column_letter(12)].width = 16
    ws.column_dimensions[openpyxl.utils.get_column_letter(13)].width = 18
    ws.column_dimensions[openpyxl.utils.get_column_letter(14)].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(16)].width = 7
    ws.column_dimensions[openpyxl.utils.get_column_letter(18)].width = 7
    ws.column_dimensions[openpyxl.utils.get_column_letter(20)].width = 7
    ws.column_dimensions[openpyxl.utils.get_column_letter(21)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(22)].width = 20
    ws.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 15
    for col in [3, 6, 7]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    for col in [5, 9]:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 13

    ws.row_dimensions[2].height = 45.75

    # Dictionary for matching bcode based on length and part
    bcode_by_length_and_part = {}
    for length, bcodes in lst_data.items():
        for bcode in bcodes:
            bcode_parts = bcode.split('-')
            if len(bcode_parts) >= 3:  # Match barcodes with 3 or more segments
                part = bcode_parts[-1]  # Use the last segment as part
                key = (length, part)
                if key not in bcode_by_length_and_part:
                    bcode_by_length_and_part[key] = []
                bcode_by_length_and_part[key].append(bcode)

    bcode_tracker = {key: 0 for key in bcode_by_length_and_part}

    # Dictionary for mapping Profile type and Material to Bar number resume
    profile_material_to_bar_number = {
        (row["Profile type resume"], row["Material resume"]): row["Bar number resume"]
        for _, row in resume_data.iterrows()
    }

    # Dictionary to store Profile type + Material combinations
    profile_material_combinations = {}

    # Fill main data with sequential bcode logic
    for i in range(len(data["Part"])):
        part_value = data["Part"][i]
        cut_off_length = data["Cut off Length"][i]

        key = (cut_off_length, part_value)
        if key in bcode_by_length_and_part:
            available_bcodes = bcode_by_length_and_part[key]
            index = bcode_tracker[key] % len(available_bcodes)
            part_value = available_bcodes[index]
            bcode_tracker[key] += 1
        else:
            part_value = data["Part"][i]

        profile_type = data["Profile type"][i]
        material = data["Material"][i]
        profile_material_key = (profile_type, material)
        if profile_material_key not in profile_material_combinations:
            profile_material_combinations[profile_material_key] = []
        profile_material_combinations[profile_material_key].append(i)

        split_profile = split_profile_type(profile_type)

        row = [
            profile_type,
            data["Bar-codenr"][i],
            data["Length bar"][i],
            material,
            data["Bar number"][i],
            data["Total length"][i],
            data["Scrap-iron"][i],
            part_value,
            cut_off_length,
            "",
            "",
            "",  # BAR-CODE for resume section (filled later)
            profile_type,  # Profile Type for resume section
            split_profile[1],  # Height
            split_profile[2],  # X
            split_profile[3],  # Width
            split_profile[4],  # X
            split_profile[5],  # Thick1
            split_profile[6],  # X
            split_profile[7],  # Thick2
            material,  # MAT2
            ""  # Bar number resume (filled later)
        ]
        ws.append(row)

    # Fill resume section
    resume_row_idx = 3
    resume_row_mapping = {}
    for (profile_type, material), indices in sorted(profile_material_combinations.items()):
        split_profile = split_profile_type(profile_type)
        bar_number_resume = profile_material_to_bar_number.get((profile_type, material), 0)
        ws.cell(row=resume_row_idx, column=13).value = profile_type
        ws.cell(row=resume_row_idx, column=14).value = split_profile[1]
        ws.cell(row=resume_row_idx, column=15).value = split_profile[2]  # X
        ws.cell(row=resume_row_idx, column=16).value = split_profile[3]
        ws.cell(row=resume_row_idx, column=17).value = split_profile[4]  # X
        ws.cell(row=resume_row_idx, column=18).value = split_profile[5]
        ws.cell(row=resume_row_idx, column=19).value = split_profile[6]  # X
        ws.cell(row=resume_row_idx, column=20).value = split_profile[7]
        ws.cell(row=resume_row_idx, column=21).value = material
        ws.cell(row=resume_row_idx, column=22).value = bar_number_resume
        resume_row_mapping[resume_row_idx] = (profile_type, material, indices)
        resume_row_idx += 1

    # Convert specific columns to numbers
    for excel_row in range(3, ws.max_row + 1):
        convert_to_number(ws, excel_row)

    # Match CSV data for BAR-CODE
    for excel_row in range(3, ws.max_row + 1):
        excel_col_a = ws.cell(row=excel_row, column=1).value
        excel_col_c = ws.cell(row=excel_row, column=3).value
        excel_col_d = ws.cell(row=excel_row, column=4).value

        excel_col_a_str = str(excel_col_a).strip() if excel_col_a is not None else ""
        excel_col_d_str = str(excel_col_d).strip() if excel_col_d is not None else ""
        excel_col_c_int = int(excel_col_c) if isinstance(excel_col_c, float) else excel_col_c
        excel_col_c_str = str(excel_col_c_int).strip() if excel_col_c_int is not None else ""

        for csv_row in range(len(csv_data)):
            if len(csv_data[csv_row]) < 3:
                continue

            csv_col_b = csv_data[csv_row][1].strip()
            csv_col_c = csv_data[csv_row][2].strip()
            csv_col_d = csv_data[csv_row][3].strip()

            if excel_col_a_str == csv_col_b and excel_col_c_str == csv_col_c and excel_col_d_str == csv_col_d:
                ws.cell(row=excel_row, column=2).value = csv_data[csv_row][0]
                convert_to_number_barcode(ws, excel_row)
                break

    # Fill BAR-CODE column in resume section
    for resume_row, (profile_type, material, indices) in resume_row_mapping.items():
        bar_codes = []
        for idx in indices:
            barcode_value = ws.cell(row=idx + 3, column=2).value
            if barcode_value is not None:
                try:
                    bar_code_int = int(float(str(barcode_value)))
                    bar_codes.append(str(bar_code_int))
                except (ValueError, TypeError):
                    bar_codes.append(str(barcode_value))
        bar_code_str = ", ".join(sorted(set(bar_codes)))
        ws.cell(row=resume_row, column=12).value = bar_code_str

    # New logic: Clear columns L to V (12 to 22) if column L (12) is empty
    for row in range(3, ws.max_row + 1):
        barcode_cell = ws.cell(row=row, column=12).value
        if barcode_cell is None or str(barcode_cell).strip() == "":
            for col in range(12, 23):  # Columns L to V (12 to 22)
                ws.cell(row=row, column=col).value = None

    # Set alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if 1 <= cell.column <= 9 or 12 <= cell.column <= 22:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            if cell.column == 1 or cell.column == 13 or cell.column == 8:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Remove duplicate rows in columns A to G
    unique_rows = set()
    for row in range(3, ws.max_row + 1):
        row_data = tuple(
            ws.cell(row=row, column=col).value for col in range(1, 8)
        )
        if row_data in unique_rows:
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.border = None
        else:
            unique_rows.add(row_data)

    # Set borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    top_thick_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thick"),
        bottom=Side(style="thin")
    )

    for cell in ws[2]:
        if 1 <= cell.column <= 9:
            cell.border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        if all(ws.cell(row=row[0].row, column=col).value not in [None, "", " "] for col in range(1, 10)):
            for cell in row:
                if 1 <= cell.column <= 9:
                    cell.border = top_thick_border
        else:
            for cell in row:
                if 1 <= cell.column <= 9 and cell.value not in [None, "", " "]:
                    cell.border = thin_border

    last_row_with_data = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=12, max_col=22):
        if any(cell.value not in [None, "", " "] for cell in row):
            last_row_with_data = row[0].row

    if last_row_with_data > 0:
        for cell in ws.iter_rows(min_row=2, max_row=last_row_with_data, min_col=12, max_col=22):
            for sub_cell in cell:
                sub_cell.border = thin_border

    for cell in ws[2]:
        if 12 <= cell.column <= 22:
            cell.border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )

    for row in range(3, ws.max_row + 1):
        ws.row_dimensions[row].height = 25

    wb.save(output_file_path)

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Convert Partlist files to Excel')
    parser.add_argument('list_file', help='Path to the .list file')
    parser.add_argument('lst_file', help='Path to the .lst file')
    parser.add_argument('csv_file', help='Path to the CSV database file')
    parser.add_argument('output_file', help='Path for the output Excel file')
    
    args = parser.parse_args()
    
    print("Processing files...")
    resume_df = process_file(args.list_file)
    convert_list_to_xlsx(args.list_file, args.lst_file, resume_df, args.csv_file, args.output_file)
    print(f"Successfully created output file: {args.output_file}")

if __name__ == "__main__":
    main()